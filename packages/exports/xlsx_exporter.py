# ContractLens — XLSX export of risk register (section 1.3)
# Generates an .xlsx file with formatted sheets using openpyxl.

from __future__ import annotations

import io
from datetime import date
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from packages.domain.schemas import LegalFinding, RiskScore, Contract


# Style constants
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Severity colors
SEVERITY_FILLS = {
    "CRITICAL": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "HIGH": PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
    "LOW": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
    "INFO": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}


def export_risk_register_xlsx(
    contract: Contract,
    findings: Sequence[LegalFinding],
    risk_score: RiskScore | None,
    output_date: date | None = None,
) -> bytes:
    """Generate an XLSX workbook bytes for the risk register.
    
    Contains one sheet per contract with formatted findings.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Risk - {contract.file_name[:25]}"

    # ── Header row ──
    headers = [
        "Finding Code", "Severity", "Title", "Description",
        "Statute Reference", "Deterministic", "Financial Impact (INR)",
        "Risk Score", "Risk Level",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # ── Data rows ──
    for row_idx, finding in enumerate(findings, start=2):
        values = [
            finding.finding_code,
            finding.severity,
            finding.title,
            finding.description,
            finding.statute_reference or "",
            "Yes" if finding.deterministic else "No",
            float(finding.financial_impact_inr or 0),
            float(risk_score.total_score) if risk_score else "",
            risk_score.level if risk_score else "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER
            # Color by severity
            if col_idx == 2:  # severity column
                fill = SEVERITY_FILLS.get(finding.severity)
                if fill:
                    cell.fill = fill
                    if finding.severity in ("CRITICAL", "HIGH"):
                        cell.font = Font(color="FFFFFF", bold=True)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 3, 60)
        ws.column_dimensions[col_letter].width = adjusted_width

    # ── Summary sheet ──
    ws2 = wb.create_sheet(title="Summary")
    ws2.cell(row=1, column=1, value="Contract ID").font = HEADER_FONT
    ws2.cell(row=1, column=2, value=str(contract.id)).border = THIN_BORDER
    ws2.cell(row=2, column=1, value="File Name").font = HEADER_FONT
    ws2.cell(row=2, column=2, value=contract.file_name).border = THIN_BORDER
    ws2.cell(row=3, column=1, value="Vendor").font = HEADER_FONT
    ws2.cell(row=3, column=2, value=contract.vendor_name or "").border = THIN_BORDER
    ws2.cell(row=4, column=1, value="Total Score").font = HEADER_FONT
    ws2.cell(row=4, column=2, value=float(risk_score.total_score) if risk_score else "").border = THIN_BORDER
    ws2.cell(row=5, column=1, value="Risk Level").font = HEADER_FONT
    ws2.cell(row=5, column=2, value=risk_score.level if risk_score else "").border = THIN_BORDER
    ws2.cell(row=6, column=1, value="Total Findings").font = HEADER_FONT
    ws2.cell(row=6, column=2, value=len(findings)).border = THIN_BORDER
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 50

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
